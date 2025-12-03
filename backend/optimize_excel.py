"""
ASML Buy Back Optimizer - Python Backend
This script reads data from Base.xlsx, saves System Recommendation data to MasterDB.xlsx,
runs the optimizer script, and updates Base.xlsx with results.
"""
from flask import Flask, request, jsonify
from flask_cors import CORS
import pandas as pd
import openpyxl
from pathlib import Path
import logging
import subprocess
import shutil
import os
from functools import wraps

app = Flask(__name__)
CORS(app, supports_credentials=True)

# Import auth service
try:
    from auth_service import AuthService
    auth_service = AuthService()
    AUTH_ENABLED = True
except ImportError:
    AUTH_ENABLED = False
    logging.warning("Authentication service not available. Running without auth.")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Path to the Excel files
# Prefer public/sample_data to keep server in sync with repo; allow override via BASE_EXCEL_PATH
BASE_PATH = Path(os.environ.get('BASE_EXCEL_PATH', '')).expanduser() if os.environ.get('BASE_EXCEL_PATH') else None
if not BASE_PATH or not BASE_PATH.exists():
    # Primary path inside repo (works on Render too)
    BASE_PATH = Path(__file__).parent.parent / 'public' / 'sample_data' / 'Base.xlsx'
if not BASE_PATH.exists():
    # Fallback to legacy backend sample_data if present
    BASE_PATH = Path(__file__).parent / 'sample_data' / 'Base.xlsx'

MASTER_DB_PATH = Path(__file__).parent / 'MasterDB.xlsx'

# Authentication decorator
def require_auth(f):
    """Decorator to require authentication for endpoints"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not AUTH_ENABLED:
            return f(*args, **kwargs)
        
        # Get session token from header or cookie
        token = request.headers.get('Authorization')
        if token and token.startswith('Bearer '):
            token = token[7:]
        elif 'session_token' in request.cookies:
            token = request.cookies.get('session_token')
        else:
            return jsonify({'status': 'error', 'message': 'Authentication required'}), 401
        
        # Validate session
        is_valid, user_data = auth_service.validate_session(token)
        if not is_valid:
            return jsonify({'status': 'error', 'message': 'Invalid or expired session'}), 401
        
        # Add user data to request context
        request.user = user_data
        return f(*args, **kwargs)
    
    return decorated_function


# ============ Authentication Endpoints ============

@app.route('/api/auth/register', methods=['POST'])
def register():
    """Register a new user account"""
    if not AUTH_ENABLED:
        return jsonify({'status': 'error', 'message': 'Authentication not available'}), 503
    
    data = request.get_json() or {}
    
    success, message, user_data = auth_service.register_user(
        username=data.get('username', ''),
        email=data.get('email', ''),
        password=data.get('password', ''),
        full_name=data.get('full_name', ''),
        company=data.get('company', '')
    )
    
    if success:
        return jsonify({
            'status': 'success',
            'message': message,
            'user': user_data
        })
    else:
        return jsonify({
            'status': 'error',
            'message': message
        }), 400


@app.route('/api/auth/login', methods=['POST'])
def login():
    """Login and create user session"""
    if not AUTH_ENABLED:
        return jsonify({'status': 'error', 'message': 'Authentication not available'}), 503
    
    data = request.get_json() or {}
    
    success, message, user_data, session_token = auth_service.login(
        username_or_email=data.get('username', ''),
        password=data.get('password', ''),
        ip_address=request.remote_addr,
        user_agent=request.headers.get('User-Agent', '')
    )
    
    if success:
        response = jsonify({
            'status': 'success',
            'message': message,
            'user': user_data,
            'session_token': session_token
        })
        # Set session cookie
        response.set_cookie('session_token', session_token, httponly=True, 
                           max_age=7*24*60*60, samesite='Lax')  # 7 days
        return response
    else:
        return jsonify({
            'status': 'error',
            'message': message
        }), 401


@app.route('/api/auth/logout', methods=['POST'])
@require_auth
def logout():
    """Logout and destroy session"""
    if not AUTH_ENABLED:
        return jsonify({'status': 'error', 'message': 'Authentication not available'}), 503
    
    token = request.headers.get('Authorization', '').replace('Bearer ', '') or \
            request.cookies.get('session_token')
    
    if token:
        auth_service.logout(token)
    
    response = jsonify({
        'status': 'success',
        'message': 'Logged out successfully'
    })
    response.set_cookie('session_token', '', expires=0)
    return response


@app.route('/api/auth/me', methods=['GET'])
@require_auth
def get_current_user():
    """Get current authenticated user info"""
    if not AUTH_ENABLED:
        return jsonify({'status': 'error', 'message': 'Authentication not available'}), 503
    
    return jsonify({
        'status': 'success',
        'user': request.user
    })


# ============ Main Application Endpoints ============


@app.route('/api/optimize', methods=['POST'])
@app.route('/optimize', methods=['POST'])  # Add route without /api prefix for nginx compatibility
# @require_auth  # Uncomment to require authentication for optimization
def optimize():
    """
    Optimization endpoint that:
    1. Copies Base.xlsx to MasterDB.xlsx
    2. Reads System Recommendation data from frontend
    3. Saves as User Input1 and User Input2 sheets in MasterDB.xlsx
    4. Runs the Python optimizer script (main_optimizer.py)
    5. Copies output sheets (out_bb, out_tot) back to Base.xlsx as OutBase and OutProfit
    6. Returns success status
    """
    try:
        logger.info(f"Starting optimization, reading from: {BASE_PATH}")
        
        # Get JSON payload and determine mode
        data = request.get_json(silent=True) or {}
        use_excel_only = bool(os.environ.get('USE_EXCEL_INPUTS') == '1' or data.get('use_excel_inputs'))

        if use_excel_only:
            logger.info("USE_EXCEL_INPUTS enabled: reading inputs directly from MasterDB.xlsx and skipping UI write.")
            # If a source Master Excel is provided, copy it first
            master_src = os.environ.get('MASTER_EXCEL_PATH')
            if master_src and Path(master_src).exists():
                logger.info(f"Copying MASTER_EXCEL_PATH from {master_src} -> {MASTER_DB_PATH}")
                shutil.copy2(master_src, MASTER_DB_PATH)
            elif not MASTER_DB_PATH.exists():
                return jsonify({
                    'status': 'error',
                    'message': 'MasterDB.xlsx not found. Provide MASTER_EXCEL_PATH or create MasterDB.xlsx with required sheets.'
                }), 400
        else:
            # Standard flow: accept UI payload and persist to MasterDB.xlsx
            if 'systemRecommendation' not in data:
                return jsonify({
                    'status': 'error',
                    'message': 'Missing systemRecommendation data in request'
                }), 400

            system_rec_data = data['systemRecommendation']
            logger.info(f"Received System Recommendation data with {len(system_rec_data)} rows")

            if len(system_rec_data) > 0:
                logger.info(f"First row sample: {system_rec_data[0]}")

            # Step 1: Copy ALL sheets from Base.xlsx to MasterDB.xlsx
            logger.info("Copying all sheets from Base.xlsx to MasterDB.xlsx...")
            import openpyxl
            wb_base = openpyxl.load_workbook(BASE_PATH)
            wb_master = openpyxl.Workbook()
            # Remove default sheet
            if 'Sheet' in wb_master.sheetnames:
                std = wb_master['Sheet']
                wb_master.remove(std)
            for sheet in wb_base.sheetnames:
                ws_base = wb_base[sheet]
                ws_master = wb_master.create_sheet(title=sheet)
                for row in ws_base.iter_rows(values_only=True):
                    ws_master.append(row)
            wb_master.save(MASTER_DB_PATH)

            # Step 2: Prepare User Input1 data from System Recommendation
            user_input1_data = []
            for row in system_rec_data:
                prob = row.get('deal_outcome_probability', 0)
                if isinstance(prob, (int, float)) and prob <= 1:
                    prob = prob * 100

                # Preserve two decimal precision for monetary values; keep counts as integers
                def to_int(v):
                    try:
                        return int(float(v))
                    except Exception:
                        return 0
                # Monetary values must be truncated to 2 decimals (no rounding up)
                # Example: 2705.64414542864 -> 2705.64 (NOT 2705.65/2706)
                def to_money(v):
                    try:
                        from decimal import Decimal, ROUND_DOWN, InvalidOperation, localcontext
                        with localcontext() as ctx:
                            ctx.prec = 28
                            d = Decimal(str(v))
                            return float(d.quantize(Decimal('0.01'), rounding=ROUND_DOWN))
                    except (InvalidOperation, Exception):
                        try:
                            # Fallback: truncate using math for non-decimal-friendly inputs
                            import math
                            f = float(v)
                            return math.floor(f * 100.0) / 100.0
                        except Exception:
                            return 0.0
                user_input1_data.append({
                    'Machine type': row.get('machine_type', ''),
                    'Offered Bundle': to_int(row.get('offered_bundle', 0)),
                    'QTC average BB price': to_money(row.get('qtc_avg_bb_price', 0)),
                    'Units in sales pipeline': to_int(row.get('units_in_sales_pipeline', 0)),
                    'Deal outcome probability': prob,
                    'Expected pipeline units': to_int(row.get('expected_pipeline_units', 0)),
                    'Units in qualified inventory': to_int(row.get('units_in_qualified_inventory', 0)),
                    'Recommended from other inventory': to_int(row.get('recommended_from_other_inventory', 0)),
                    'Recommended Buy for 12 M': to_int(row.get('recommended_buy_12m', 0)),
                    'Required margin (%)': to_int(row.get('required_margin', 0)),
                    'Recommended BB Price on Bundle (K)': to_money(row.get('recommended_bb_price_on_bundle', 0))
                })

            df_input1 = pd.DataFrame(user_input1_data)
            logger.info(f"Created User Input1 with {len(df_input1)} rows and {len(df_input1.columns)} columns")
            logger.info(f"Columns: {list(df_input1.columns)}")
            if len(df_input1) > 0:
                logger.info(f"Sample first row: {df_input1.iloc[0].to_dict()}")

            # Step 3: Prepare User Input2 data - ensure all required metrics exist
            max_buyback_data = data.get('maxBuyback', [])
            
            # Define all required metrics that bb_recom expects
            required_metrics = [
                'Refurbishment',
                'Harvesting - Module',
                'Harvesting - Parts',
                'EOL',
                'Total Without Scrap',
                'Scrap',
                'Total With Scrap'
            ]
            
            # Build a dictionary from incoming data
            metrics_dict = {}
            if max_buyback_data and len(max_buyback_data) > 0:
                for row in max_buyback_data:
                    metric = str(row.get('metric', '')).strip()
                    required_margin_val = row.get('required_margin', 0)
                    try:
                        required_margin = int(float(required_margin_val)) if required_margin_val not in [None, ''] else 0
                    except Exception:
                        required_margin = 0
                    metrics_dict[metric] = {
                        'Metric': metric,
                        'Max_BBB_Valuation': int(float(row.get('valuation', 0))) if row.get('valuation', 0) not in [None, ''] else 0,
                        'Required Margin': required_margin
                    }
            
            # Ensure all required metrics exist with defaults
            for metric in required_metrics:
                if metric not in metrics_dict:
                    default_margin = 40  # Default 40% for all metrics
                    if metric == 'Refurbishment':
                        default_margin = int(float(data.get('refurbishmentMargin', 40)))
                    elif metric == 'Harvesting - Module':
                        default_margin = int(float(data.get('harvestingModuleMargin', 40)))
                    elif metric == 'Harvesting - Parts':
                        default_margin = int(float(data.get('harvestingPartsMargin', 40)))
                    elif metric == 'EOL':
                        default_margin = int(float(data.get('eolMargin', 40)))
                    elif metric == 'Total Without Scrap':
                        default_margin = int(float(data.get('totalWithoutScrapMargin', 40)))
                    elif metric == 'Total With Scrap':
                        default_margin = int(float(data.get('totalWithScrapMargin', 40)))
                    elif metric == 'Scrap':
                        default_margin = 0  # Scrap has no margin requirement
                    metrics_dict[metric] = {
                        'Metric': metric,
                        'Max_BBB_Valuation': 0,
                        'Required Margin': default_margin
                    }
            
            # Build list in the order of required_metrics to maintain consistent ordering
            user_input2_data = [metrics_dict[metric] for metric in required_metrics if metric in metrics_dict]

            df_input2 = pd.DataFrame(user_input2_data)
            logger.info(f"Created User Input2 with {len(df_input2)} rows")

            # Step 3.5: Save Systems, Modules, Parts demand data to MasterDB.xlsx
            systems_data = data.get('systems', [])
            modules_data = data.get('modules', [])
            parts_data = data.get('parts', [])

            systems_list = [{
                'System': row.get('item', ''),
                'Demand_12M': int(float(row.get('demand_12m', 0))),
                'Demand_24M': int(float(row.get('demand_24m', 0))),
                'Qinventory_12M': int(float(row.get('finished_12m', 0))),
                'Qinventory_24M': int(float(row.get('finished_24m', 0)))
            } for row in systems_data]

            modules_list = [{
                'Module': row.get('item', ''),
                'Demand_12M': int(float(row.get('demand_12m', 0))),
                'Demand_24M': int(float(row.get('demand_24m', 0))),
                'Qinventory_12M': int(float(row.get('finished_12m', 0))),
                'Qinventory_24M': int(float(row.get('finished_24m', 0)))
            } for row in modules_data]

            parts_list = [{
                'Module': row.get('item', ''),
                'Demand_12M': int(float(row.get('demand_12m', 0))),
                'Demand_24M': int(float(row.get('demand_24m', 0))),
                'Qinventory_12M': int(float(row.get('finished_12m', 0))),
                'Qinventory_24M': int(float(row.get('finished_24m', 0)))
            } for row in parts_data]

            df_systems = pd.DataFrame(systems_list)
            df_modules = pd.DataFrame(modules_list)
            df_parts = pd.DataFrame(parts_list)

            with pd.ExcelWriter(MASTER_DB_PATH, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df_input1.to_excel(writer, sheet_name='User Input1', index=False)
                df_input2.to_excel(writer, sheet_name='User Input2', index=False)
                if len(df_systems) > 0:
                    df_systems.to_excel(writer, sheet_name='Systems', index=False)
                if len(df_modules) > 0:
                    df_modules.to_excel(writer, sheet_name='Modules', index=False)
                if len(df_parts) > 0:
                    df_parts.to_excel(writer, sheet_name='Parts', index=False)
            logger.info("Saved all data to MasterDB.xlsx")
        
        # Step 5: Run the optimizer script
        logger.info("Running optimizer script...")
        
        # Detect platform and set correct Python path
        import platform
        import sys
        
        optimizer_script = Path(__file__).parent / 'main_optimizer.py'
        
        # On Windows, use shell=True to avoid Win32 application error
        if platform.system() == 'Windows':
            logger.info(f"Running on Windows with Python: {sys.executable}")
            logger.info(f"Optimizer script: {optimizer_script}")
            
            result = subprocess.run(
                f'"{sys.executable}" "{optimizer_script}"',
                shell=True,
                cwd=str(Path(__file__).parent),
                capture_output=True,
                text=True,
                timeout=60
            )
        else:
            # Unix-like systems (Mac, Linux) - use list format
            python_path = sys.executable
            venv_python = Path(__file__).parent.parent / '.venv' / 'bin' / 'python'
            if venv_python.exists():
                python_path = str(venv_python)
            
            logger.info(f"Running on {platform.system()} with Python: {python_path}")
            logger.info(f"Optimizer script: {optimizer_script}")
            
            result = subprocess.run(
                [python_path, str(optimizer_script)],
                cwd=str(Path(__file__).parent),
                capture_output=True,
                text=True,
                timeout=60
            )
        
        if result.returncode != 0:
            logger.error(f"Optimizer script failed: {result.stderr}")
            return jsonify({
                'status': 'error',
                'message': f'Optimizer script failed: {result.stderr}'
            }), 500
        
        logger.info(f"Optimizer output: {result.stdout}")
        
        # Step 6: Read output sheets from MasterDB.xlsx (no changes to Base.xlsx)
        logger.info("Reading output sheets from MasterDB.xlsx...")
        
        # Close and reopen the workbook to ensure optimizer changes are visible
        import time
        time.sleep(0.5)  # Give file system a moment to sync
        
        master_wb = openpyxl.load_workbook(MASTER_DB_PATH)
        
        # Debug: Log all available sheets
        logger.info(f"All sheets in MasterDB.xlsx: {master_wb.sheetnames}")
        
        # Find the generated output sheets (look for out_bb variations: out_bb, out_bb1, Out_BB, etc.)
        out_bb_sheet_name = None
        out_tot_sheet_name = None
        
        for sheet in master_wb.sheetnames:
            if sheet.lower().startswith('out_bb') or sheet.lower() == 'out_bb':
                out_bb_sheet_name = sheet
            if sheet.lower().startswith('out_tot') or sheet.lower() == 'out_tot':
                out_tot_sheet_name = sheet
        
        logger.info(f"Found output sheets - OutBase: {out_bb_sheet_name}, OutProfit: {out_tot_sheet_name}")
        
        # Read User Input1 data to return to frontend (System Recommendation green columns)
        user_input1_data = []
        if 'User Input1' in master_wb.sheetnames:
            input1_ws = master_wb['User Input1']
            # Read header row
            headers = [cell.value for cell in input1_ws[1]]
            # Read data rows
            for row in input1_ws.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for idx, value in enumerate(row):
                    if idx < len(headers) and headers[idx]:
                        row_dict[headers[idx]] = value
                user_input1_data.append(row_dict)
            logger.info(f"Read {len(user_input1_data)} rows from User Input1")
        
        # Read User Input2 data to return to frontend (Max Buy Back required margins)
        user_input2_data = []
        if 'User Input2' in master_wb.sheetnames:
            input2_ws = master_wb['User Input2']
            # Read header row
            headers = [cell.value for cell in input2_ws[1]]
            # Read data rows
            for row in input2_ws.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for idx, value in enumerate(row):
                    if idx < len(headers) and headers[idx]:
                        row_dict[headers[idx]] = value
                user_input2_data.append(row_dict)
            logger.info(f"Read {len(user_input2_data)} rows from User Input2")
        
        # Read OutBase data to return to frontend (System Recommendation yellow columns)
        # Use the original generated sheet name (out_bb1, out_bb, etc.)
        outbase_data = []
        if out_bb_sheet_name and out_bb_sheet_name in master_wb.sheetnames:
            outbase_ws = master_wb[out_bb_sheet_name]
            # Read header row
            headers = [cell.value for cell in outbase_ws[1]]
            # Read data rows
            for row in outbase_ws.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for idx, value in enumerate(row):
                    if idx < len(headers) and headers[idx]:
                        # Preserve numeric precision (round to 2 decimals only for floats)
                        if isinstance(value, (int, float)) and not isinstance(value, bool):
                            if isinstance(value, float):
                                row_dict[headers[idx]] = round(value, 2)
                            else:
                                row_dict[headers[idx]] = value
                        else:
                            row_dict[headers[idx]] = value
                outbase_data.append(row_dict)
            logger.info(f"Read {len(outbase_data)} rows from {out_bb_sheet_name}")
        
        # Read OutProfit data to return to frontend (Max Buy Back and Expected Profit tables)
        # Use the original generated sheet name (out_tot1, out_tot, etc.)
        outprofit_data = []
        if out_tot_sheet_name and out_tot_sheet_name in master_wb.sheetnames:
            outprofit_ws = master_wb[out_tot_sheet_name]
            # Read header row
            headers = [cell.value for cell in outprofit_ws[1]]
            # Read data rows
            for row in outprofit_ws.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for idx, value in enumerate(row):
                    if idx < len(headers) and headers[idx]:
                        # Preserve precision for numeric values (2 decimals for floats)
                        if isinstance(value, (int, float)) and not isinstance(value, bool):
                            if isinstance(value, float):
                                row_dict[headers[idx]] = round(value, 2)
                            else:
                                row_dict[headers[idx]] = value
                        else:
                            row_dict[headers[idx]] = value
                outprofit_data.append(row_dict)
            logger.info(f"Read {len(outprofit_data)} rows from {out_tot_sheet_name}")
        
        # Optional warnings sheet
        warnings_list = []
        if 'warnings' in master_wb.sheetnames:
            warn_ws = master_wb['warnings']
            headers = [cell.value for cell in warn_ws[1]]
            for row in warn_ws.iter_rows(min_row=2, values_only=True):
                w = {}
                for idx, value in enumerate(row):
                    if idx < len(headers) and headers[idx]:
                        w[headers[idx]] = value
                warnings_list.append(w)

        master_wb.close()
        
        logger.info(f"Saved results to: {MASTER_DB_PATH}")
        
        response = {
            'status': 'success',
            'message': 'Optimization completed successfully. Results saved to MasterDB.xlsx',
            'output_file': 'MasterDB.xlsx',
            'outbase_data': outbase_data,
            'outprofit_data': outprofit_data,
            'warnings': warnings_list
        }

        # Only include input echoes if we wrote them from UI
        if not use_excel_only:
            response.update({
                'sheets_updated': ['User Input1', 'User Input2', 'OutBase', 'OutProfit'],
                'user_input1_data': user_input1_data,
                'user_input2_data': user_input2_data,
            })

        return jsonify(response)
        
    except subprocess.TimeoutExpired:
        logger.error("Optimizer script timed out")
        return jsonify({
            'status': 'error',
            'message': 'Optimization timed out after 60 seconds'
        }), 500
    except Exception as e:
        logger.error(f"Optimization failed: {str(e)}", exc_info=True)
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

@app.route('/health', methods=['GET'])
def health():
    """Health check endpoint"""
    return jsonify({'status': 'healthy'})

@app.route('/', methods=['GET'])
def home():
    """Root endpoint"""
    return jsonify({
        'status': 'running',
        'service': 'ASML Buy Back Optimiser Backend',
        'endpoints': {
            'health': '/health',
            'optimize': '/api/optimize (POST)'
        }
    })

if __name__ == '__main__':
    logger.info(f"Starting optimizer server, Excel file: {BASE_PATH}")
    # Use 0.0.0.0 to accept connections from Docker network
    # Use port 5001 to match docker-compose configuration
    port = int(os.environ.get('PORT', 5001))
    app.run('0.0.0.0', port=port, debug=False)
